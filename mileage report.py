import os, re, csv, io
from collections import defaultdict

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Mileage Report Generator", page_icon="🚗", layout="centered")

st.title("🚗 Mileage Report Generator")
st.markdown("Upload all your CSV mileage files below, choose a report type, and download the Excel summary.")


def parse_csv_content(content: str):
    lines = [l.rstrip() for l in content.strip().split("\n")]
    header_match = re.search(
        r"Name:\s*(.+?),\s*Member:\s*(.+?),\s*Month/Year:\s*(.+)", lines[0]
    )
    if not header_match:
        return None

    staff_name  = header_match.group(1).strip()
    member_name = header_match.group(2).strip()
    month_year  = header_match.group(3).strip()

    total_km = 0.0
    total_parking = 0.0

    for line in lines:
        km_match = re.match(r"Total Kilometers,\s*([\d.]+)", line)
        if km_match:
            total_km = float(km_match.group(1))

    data_started = False
    for line in lines:
        if line.startswith("Date,"):
            data_started = True
            continue
        if data_started:
            if line == "" or line.startswith(("Total", "Approved", "Month Billed")):
                continue
            try:
                row = next(csv.reader(io.StringIO(line)))
                if len(row) >= 6:
                    p = row[5].strip()
                    total_parking += float(p) if p not in ("", "null") else 0.0
            except Exception:
                pass

    return {"staff": staff_name, "member": member_name,
            "month_year": month_year,
            "total_km": round(total_km, 2),
            "total_parking": round(total_parking, 2)}


def make_styles():
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    return border, center, left


def build_per_member(records) -> bytes:
    members = defaultdict(list)
    for r in records:
        members[r["member"]].append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "By Member"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    member_fill = PatternFill("solid", fgColor="2E5D9E")
    label_fill  = PatternFill("solid", fgColor="D9E1F2")
    data_fill   = PatternFill("solid", fgColor="EEF2FB")
    border, center, left = make_styles()

    row_cursor = 1
    for idx, (member_name, staff_list) in enumerate(sorted(members.items()), start=1):
        num_staff = len(staff_list)
        member_total_km      = round(sum(s["total_km"]      for s in staff_list), 2)
        member_total_parking = round(sum(s["total_parking"] for s in staff_list), 2)

        ws.merge_cells(start_row=row_cursor, start_column=1,
                       end_row=row_cursor, end_column=1 + num_staff)
        cell = ws.cell(row=row_cursor, column=1,
                       value=f"Member {idx}: {member_name}  —  {member_total_km} km  |  ${member_total_parking:.2f} parking")
        cell.font = header_font; cell.fill = member_fill
        cell.alignment = center; cell.border = border
        row_cursor += 1

        ws.cell(row=row_cursor, column=1, value="").border = border
        for col, staff in enumerate(staff_list, start=2):
            c = ws.cell(row=row_cursor, column=col, value=staff["staff"])
            c.font = Font(bold=True, size=10)
            c.fill = label_fill; c.alignment = center; c.border = border
        row_cursor += 1

        lbl = ws.cell(row=row_cursor, column=1, value="Total KMs")
        lbl.font = Font(bold=True); lbl.fill = label_fill
        lbl.alignment = left; lbl.border = border
        for col, staff in enumerate(staff_list, start=2):
            c = ws.cell(row=row_cursor, column=col, value=staff["total_km"])
            c.fill = data_fill; c.alignment = center; c.border = border
        row_cursor += 1

        lbl = ws.cell(row=row_cursor, column=1, value="Parking Expense ($)")
        lbl.font = Font(bold=True); lbl.fill = label_fill
        lbl.alignment = left; lbl.border = border
        for col, staff in enumerate(staff_list, start=2):
            c = ws.cell(row=row_cursor, column=col, value=staff["total_parking"])
            c.fill = data_fill; c.alignment = center; c.border = border
        row_cursor += 2

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 18)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_per_staff(records) -> bytes:
    staff_map = defaultdict(list)
    for r in records:
        staff_map[r["staff"]].append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "By Staff"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    staff_fill  = PatternFill("solid", fgColor="2E5D9E")
    col_fill    = PatternFill("solid", fgColor="4472C4")
    label_fill  = PatternFill("solid", fgColor="D9E1F2")
    data_fill   = PatternFill("solid", fgColor="EEF2FB")
    border, center, left = make_styles()

    row_cursor = 1
    for idx, (staff_name, members) in enumerate(sorted(staff_map.items()), start=1):
        staff_total_km      = round(sum(m["total_km"]      for m in members), 2)
        staff_total_parking = round(sum(m["total_parking"] for m in members), 2)

        ws.merge_cells(start_row=row_cursor, start_column=1,
                       end_row=row_cursor, end_column=3)
        cell = ws.cell(row=row_cursor, column=1,
                       value=f"Staff {idx}: {staff_name}  —  {staff_total_km} km  |  ${staff_total_parking:.2f} parking")
        cell.font = header_font; cell.fill = staff_fill
        cell.alignment = center; cell.border = border
        row_cursor += 1

        for col, h in enumerate(["Member", "Total KMs", "Parking Expense ($)"], start=1):
            c = ws.cell(row=row_cursor, column=col, value=h)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = col_fill; c.alignment = center; c.border = border
        row_cursor += 1

        for member in sorted(members, key=lambda x: x["member"]):
            ml = ws.cell(row=row_cursor, column=1, value=member["member"])
            ml.fill = label_fill; ml.alignment = left; ml.border = border

            km = ws.cell(row=row_cursor, column=2, value=member["total_km"])
            km.fill = data_fill; km.alignment = center; km.border = border

            pk = ws.cell(row=row_cursor, column=3, value=member["total_parking"])
            pk.fill = data_fill; pk.alignment = center; pk.border = border
            row_cursor += 1

        for col, val in enumerate([("TOTAL", left, label_fill),
                                    (staff_total_km, center, label_fill),
                                    (staff_total_parking, center, label_fill)], start=1):
            v, align, fill = val
            c = ws.cell(row=row_cursor, column=col, value=v)
            c.font = Font(bold=True); c.fill = fill
            c.alignment = align; c.border = border
        row_cursor += 2

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 18)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


uploaded_files = st.file_uploader(
    "Drop your CSV mileage files here",
    type="csv",
    accept_multiple_files=True
)

if uploaded_files:
    records = []
    failed  = []
    for f in uploaded_files:
        content = f.read().decode("utf-8")
        result  = parse_csv_content(content)
        if result:
            records.append(result)
        else:
            failed.append(f.name)

    st.success(f"✅ Parsed **{len(records)}** file(s) successfully.")
    if failed:
        st.warning(f"⚠️ Could not parse: {', '.join(failed)}")

    if records:
        df = pd.DataFrame(records)[["staff", "member", "month_year", "total_km", "total_parking"]]
        df.columns = ["Staff", "Member", "Month/Year", "Total KMs", "Parking ($)"]
        st.dataframe(df, use_container_width=True)

        st.divider()
        st.subheader("Download Reports")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**📋 Per Member**")
            st.caption("One block per member, staff listed as columns")
            per_member_bytes = build_per_member(records)
            st.download_button(
                label="⬇️ Download Per-Member Report",
                data=per_member_bytes,
                file_name="mileage_by_member.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col2:
            st.markdown("**👤 Per Staff**")
            st.caption("One block per staff, members listed as rows")
            per_staff_bytes = build_per_staff(records)
            st.download_button(
                label="⬇️ Download Per-Staff Report",
                data=per_staff_bytes,
                file_name="mileage_by_staff.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
else:
    st.info("👆 Upload one or more CSV files to get started.")
